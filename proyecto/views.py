# -*- encoding: utf-8 -*-
from unittest import result

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, render_to_response
from django.http import HttpResponseRedirect,HttpResponse
from django.template import RequestContext
from django.contrib.auth.hashers import make_password
from .models import  Usuario
from .models import  Persona
from .models import  docente,Alumno2,Grado,Sisben,Estrato,Municipios,Veredas,Universidades,Colegios,CentrosSena,FormularioEstimulos,Sedes,\
CentrosSena,Horas,Fechas,Traza,Grupos2,Grado,Inst_educativas2,SedeInstitucion2,Video
from validators import Validator, FormLoginValidator
from django.contrib import auth
import xhtml2pdf.pisa as pisa
from StringIO import StringIO
from django.template.loader import render_to_string
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.core import serializers
import json
from django.db.models import Q
from openpyxl import Workbook
from django.contrib.auth.decorators import permission_required
# Create your views here.
def login(request):
     """view del login
    """
    #Verificamos que los datos lleguen por el methodo POST

     if request.method == 'POST':
        #Cargamos el formulario (ver forms.py con los datos del POST)
        validator = FormLoginValidator(request.POST)
        #formulario = LoginForm(data = request.POST)
        #Verificamos que los datos esten correctos segun su estructura

        if validator.is_valid():
            # Capturamos las variables que llegan por POST

            auth.login(request, validator.acceso) # Crear una sesion
            return HttpResponseRedirect(request.POST['next'])
        else:
            return render_to_response('login.html', {'error': validator.getMessage() } , context_instance = RequestContext(request))
     next = '/portada'
     if 'next' in request.GET:
         next = request.GET['next']

     return render_to_response('login.html', {'next': next} ,context_instance = RequestContext(request))

@permission_required('Usuario.can_add_Usuario', login_url = '/login')
def registro(request):
    if request.method == 'POST':
        validators = Validator(request.POST)
        validators.required = ['nombre','user','password']
        if validators.is_valid():
            usuario = Usuario()
            usuario.first_name = request.POST['nombre']
            usuario.last_name = request.POST['apellido']
            usuario.username = request.POST['user']
            usuario.email = request.POST['email']
            usuario.password = make_password(request.POST['password'])
            usuario.is_active = True
            usuario.save()
            return render_to_response('portada.html',context_instance = RequestContext(request))
        else:
            return render_to_response('registrarse.html', {'error': validator.getMessage() } , context_instance = RequestContext(request))
    return render_to_response('formulario.html',context_instance = RequestContext(request))

@login_required(login_url = "/login")
def persona(request):
    if request.method == 'POST':
        validators = Validator(request.POST)
        validators.required = ['nombre','apellido','documento','email','sexo']
        if validators.is_valid():
            alumno = Alumno()
            alumno.nombres = request.POST['nombre']
            alumno.apellidos = request.POST['apellido']
            alumno.documento = request.POST['documento']
            alumno.email = request.POST['email']
            alumno.sexo = request.POST['sexo']
            alumno.save()
            print(alumno.nombres)
        else:
            return render_to_response('registrarse.html', {'error': validator.getMessage() } , context_instance = RequestContext(request))
    return render_to_response('formulario.html',context_instance = RequestContext(request))

@login_required(login_url = "/login")
def inicio(request):
    return render_to_response('inicio.html',context_instance = RequestContext(request))

@login_required(login_url = "/login")
def configuraMatricula(request):
    return render_to_response('configuraMatriculas.html',context_instance = RequestContext(request))

@login_required(login_url = "/login")
def configuraInstitucion(request):
    instituciones = Inst_educativas2.objects.all()
    if request.method == 'POST':
        institucion = Inst_educativas2()
        institucion.nombre = request.POST['institucion']
        institucion.save()
        return render_to_response('configuraInstitucion.html',{'instituciones':instituciones}, context_instance = RequestContext(request))
    return render_to_response('configuraInstitucion.html',{'instituciones':instituciones}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def editarInstitucion(request, pk):
    institucion = get_object_or_404(Inst_educativas2, pk=pk)
    if request.method == 'POST':
        Inst_educativa = get_object_or_404(Inst_educativas2, pk=pk)
        Inst_educativa.nombre = request.POST['institucion']
        Inst_educativa.direccion = request.POST['direccion']
        Inst_educativa.save()
        instituciones = Inst_educativas2.objects.all()
        return render_to_response('configuraInstitucion.html',{'instituciones':instituciones}, context_instance = RequestContext(request))
    return render_to_response('editarInstitucion.html',{'institucion':institucion}, context_instance = RequestContext(request))

@permission_required('Inst_educativas2.can_delete_Inst_educativas2', login_url = '/login')
def eliminarInstitucion(request, pk):
    institucion = get_object_or_404(Inst_educativas2, pk=pk)
    if request.method == 'POST':
        institucion.delete()
        instituciones = Inst_educativas2.objects.all()
        return render_to_response('configuraInstitucion.html',{'instituciones':instituciones}, context_instance = RequestContext(request))
    return render_to_response('eliminarInstitucion.html',{'institucion':institucion}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def configuraGrado(request):
    grados = Grado.objects.all()
    if request.method == 'POST':
        grado = Grado()
        grado.grado = request.POST['grado']
        grado.save()
        return render_to_response('configuraGrado.html',{'grados':grados}, context_instance = RequestContext(request))
    return render_to_response('configuraGrado.html',{'grados':grados}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def editarGrado(request, pk):
    grado = get_object_or_404(Grado, pk=pk)
    if request.method == 'POST':
        grado = get_object_or_404(Grado, pk=pk)
        grado.grado = request.POST['grado']
        grado.save()
        grados = Grado.objects.all()
        return render_to_response('configuraGrado.html',{'grados':grados}, context_instance = RequestContext(request))
    return render_to_response('editarGrado.html',{'grado':grado}, context_instance = RequestContext(request))

@permission_required('Grado.can_delete_Grado', login_url = '/login')
def eliminarGrado(request, pk):
    grado = get_object_or_404(Grado, pk=pk)
    if request.method == 'POST':
        grado.delete()
        grados = Grado.objects.all()
        return render_to_response('configuraGrado.html',{'grados':grados}, context_instance = RequestContext(request))
    return render_to_response('eliminarGrado.html',{'grado':grado}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def configuraSede(request):
    sedes = SedeInstitucion2.objects.all()
    instituciones = Inst_educativas2.objects.all()
    if request.method == 'POST':
        sede = SedeInstitucion2()
        sede.sede = request.POST['sede']
        sede.institucion_id = request.POST['institucion']
        sede.save()
        return render_to_response('configuraSede.html',{'sedes':sedes, 'instituciones':instituciones}, context_instance = RequestContext(request))
    return render_to_response('configuraSede.html',{'sedes':sedes, 'instituciones':instituciones}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def editarSede(request, pk):
    sede = get_object_or_404(SedeInstitucion2, pk=pk)
    instituciones = Inst_educativas2.objects.all()
    if request.method == 'POST':
        sede = get_object_or_404(SedeInstitucion2, pk=pk)
        sede.sede = request.POST['sede']
        sede.institucion_id = request.POST['institucion']
        sede.save()
        sedes = SedeInstitucion2.objects.all()
        return render_to_response('configuraSede.html',{'sedes':sedes, 'instituciones':instituciones}, context_instance = RequestContext(request))
    return render_to_response('editarSede.html',{'sede':sede, 'instituciones':instituciones}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
@permission_required('SedeInstitucion2.can_delete_SedeInstitucion2', login_url = '/login')
def eliminarSede(request, pk):
    sede = get_object_or_404(SedeInstitucion2, pk=pk)
    if request.method == 'POST':
        sede.delete()
        sedes = SedeInstitucion2.objects.all()
        return render_to_response('configuraSede.html',{'sedes':sedes}, context_instance = RequestContext(request))
    return render_to_response('eliminarSede.html',{'sede':sede}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def configuraGrupo(request):
    grupos = Grupos2.objects.all()
    grados = Grado.objects.all()
    if request.method == 'POST':
        grupo = Grupos2()
        grupo.grupo = request.POST['grupo']
        grupo.grado_id = request.POST['grado']
        grupo.save()
        return render_to_response('configuraGrupo.html',{'grupos':grupos, 'grados':grados}, context_instance = RequestContext(request))
    return render_to_response('configuraGrupo.html',{'grupos':grupos, 'grados':grados}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def editarGrupo(request, pk):
    grupo = get_object_or_404(Grupos2, pk=pk)
    grados = Grado.objects.all()
    if request.method == 'POST':
        grupo = get_object_or_404(Grupos2, pk=pk)
        grupo.grupo = request.POST['grupo']
        grupo.grado_id = request.POST['grado']
        grupo.save()
        grupos = Grupos2.objects.all()
        return render_to_response('configuraGrupo.html',{'grupos':grupos, 'grados':grados}, context_instance = RequestContext(request))
    return render_to_response('editarGrupo.html',{'grupo':grupo, 'grados':grados}, context_instance = RequestContext(request))

@permission_required('Grupos2.can_delete_Grupos2', login_url = '/login')
def eliminarGrupo(request, pk):
    grupo = get_object_or_404(Grupos2, pk=pk)
    if request.method == 'POST':
        grupo.delete()
        grupos = Grupos2.objects.all()
        return render_to_response('configuraGrupo.html',{'grupos':grupos}, context_instance = RequestContext(request))
    return render_to_response('eliminarGrupo.html',{'grupo':grupo}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def busquedas2(request):
    return render_to_response('formulario_exitoso.html',context_instance = RequestContext(request))

@login_required(login_url = "/login")
def Estatico(request):
    return render_to_response('estatico.html',context_instance = RequestContext(request))

@login_required(login_url = "/login")
def InformeVarios(request):
    return render_to_response('reportesVarios.html',context_instance = RequestContext(request))

@login_required(login_url = "/login")
def informeVca(request):
    fechas = Fechas.objects.order_by('fecha','hora')
    wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'

        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda

        #Creamos los encabezados desde la celda B3 hasta la E3

    ws['b1'] = 'HORA'
    ws['c1'] = 'FECHA'
    ws['d1'] = 'NOMBRE'
    ws['e1'] = 'APELLIDO'
    ws['f1'] = 'DOCUMENTO'
    cont=4
    #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for fecha in fechas:
        ws.cell(row=cont,column=2).value = fecha.hora.hora
        ws.cell(row=cont,column=3).value = fecha.fecha
        ws.cell(row=cont,column=4).value = fecha.nombre.nombre
        ws.cell(row=cont,column=5).value = fecha.nombre.apellido
        ws.cell(row=cont,column=6).value = fecha.nombre.documento
        cont = cont + 1

        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas

        #Establecemos el nombre del archivo
    nombre_archivo ="ReporteAgendaEntrevistaExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

@login_required(login_url = "/login")
def excelAgendados(request):
    fechas = Fechas.objects.order_by('fecha','hora')
    wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        #Creamos los encabezados desde la celda B3 hasta la E3
    ws['b1'] = 'nombre'
    ws['c1'] = 'apellido'
    ws['d1'] = 'tipoDocumento'
    ws['e1'] = 'documento'
    ws['f1'] = 'documentoOtro'
    ws['g1'] = 'email'
    ws['h1'] = 'celular'
    ws['i1'] = 'genero'
    ws['j1'] = 'edad'
    ws['k1'] = 'estadoCivil'
    ws['l1'] = 'cabezaFamilia'
    ws['m1'] = 'hijos'
    ws['n1'] = 'pertenencia'
    ws['o1'] = 'pertenenciaOtro'
    ws['p1'] = 'direccionEstudiante'
    ws['q1'] = 'ciudadEstudiante'
    ws['r1'] = 'barrioEstudiante'
    ws['s1'] = 'ubicacionEstudiante'
    ws['t1'] = 'vivienda'
    ws['u1'] = 'estrato'
    ws['v1'] = 'sisben'
    ws['w1'] = 'nucleoFamiliar'
    ws['x1'] = 'direccionPadres'
    ws['y1'] = 'ciudadPadres'
    ws['z1'] = 'barrioPadres'
    ws['aa1'] = 'ubicacionPadres'
    ws['ab1'] = 'serviciosPublicos'
    ws['ac1'] = 'dependenciaEconomica'
    ws['ad1'] = 'dependenciaOtro'
    ws['ae1'] = 'ingresosFamilia'
    ws['af1'] = 'colegio'
    ws['ag1'] = 'oficialPrivado'
    ws['ah1'] = 'añoTerminacion'
    ws['ai1'] = 'actaGrado'
    ws['aj1'] = 'universidad'
    ws['ak1'] = 'centrosSena'
    ws['al1'] = 'sede'
    ws['am1'] = 'acreditacionUniversidad'
    ws['an1'] = 'carrera'
    ws['ao1'] = 'tecnicaTecnologica'
    ws['ap1'] = 'acreditacionCarrera'
    ws['aq1'] = 'cicloEducacion'
    ws['ar1'] = 'semestreEnCurso'
    ws['as1'] = 'modalidad'
    ws['at1'] = 'modalidadOtro'
    ws['au1'] = 'propedeutico'
    ws['av1'] = 'propedeuticoOtro'
    ws['aw1'] = 'valorMatricula'
    ws['ax1'] = 'beneficiarioprogramas'
    ws['ay1'] = 'programaOtro'
    ws['az1'] = 'incentivoEducativo'
    ws['ba1'] = 'cuantasVeces'
    ws['bb1'] = 'extraAcademicas'
    ws['bc1'] = 'extraAcademicaPaga'
    ws['bd1'] = 'parentesco1'
    ws['be1'] = 'parentesco2'
    ws['bf1'] = 'parentesco3'
    ws['bg1'] = 'parentesco4'
    ws['bh1'] = 'parentesco5'
    ws['bi1'] = 'parentesco6'
    ws['bj1'] = 'parentesco7'
    cont=4
    #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for fecha in fechas:
        ws.cell(row=cont,column=2).value = fecha.nombre.nombre
        ws.cell(row=cont,column=3).value = fecha.nombre.apellido
        ws.cell(row=cont,column=4).value = fecha.nombre.tipoDocumento
        ws.cell(row=cont,column=5).value = fecha.nombre.documento
        ws.cell(row=cont,column=6).value = fecha.nombre.documentoOtro
        ws.cell(row=cont,column=7).value = fecha.nombre.email
        ws.cell(row=cont,column=8).value = fecha.nombre.celular
        ws.cell(row=cont,column=9).value = fecha.nombre.genero
        ws.cell(row=cont,column=10).value = fecha.nombre.edad
        ws.cell(row=cont,column=11).value = fecha.nombre.estadoCivil
        ws.cell(row=cont,column=12).value = fecha.nombre.cabezaFamilia
        ws.cell(row=cont,column=13).value = fecha.nombre.hijos
        ws.cell(row=cont,column=14).value = fecha.nombre.pertenencia
        ws.cell(row=cont,column=15).value = fecha.nombre.pertenenciaOtro
        ws.cell(row=cont,column=16).value = fecha.nombre.direccionEstudiante
        ws.cell(row=cont,column=17).value = fecha.nombre.ciudadEstudiante
        ws.cell(row=cont,column=18).value = fecha.nombre.barrioEstudiante
        ws.cell(row=cont,column=19).value = fecha.nombre.ubicacionEstudiante
        ws.cell(row=cont,column=20).value = fecha.nombre.vivienda
        ws.cell(row=cont,column=21).value = fecha.nombre.estrato
        ws.cell(row=cont,column=22).value = fecha.nombre.sisben
        ws.cell(row=cont,column=23).value = fecha.nombre.nucleoFamiliar
        ws.cell(row=cont,column=24).value = fecha.nombre.direccionPadres
        ws.cell(row=cont,column=25).value = fecha.nombre.ciudadPadres
        ws.cell(row=cont,column=26).value = fecha.nombre.barrioPadres
        ws.cell(row=cont,column=27).value = fecha.nombre.ubicacionPadres
        ws.cell(row=cont,column=28).value = fecha.nombre.serviciosPublicos
        ws.cell(row=cont,column=29).value = fecha.nombre.dependenciaEconomica
        ws.cell(row=cont,column=30).value = fecha.nombre.dependenciaOtro
        ws.cell(row=cont,column=31).value = fecha.nombre.ingresosFamilia
        ws.cell(row=cont,column=32).value = fecha.nombre.colegio
        ws.cell(row=cont,column=33).value = fecha.nombre.oficialPrivado
        ws.cell(row=cont,column=34).value = fecha.nombre.anoTerminacion
        ws.cell(row=cont,column=35).value = fecha.nombre.actaGrado
        ws.cell(row=cont,column=36).value = fecha.nombre.universidad
        ws.cell(row=cont,column=37).value = fecha.nombre.centrosSena
        ws.cell(row=cont,column=38).value = fecha.nombre.sede
        ws.cell(row=cont,column=39).value = fecha.nombre.acreditacionUniversidad
        ws.cell(row=cont,column=40).value = fecha.nombre.carrera
        ws.cell(row=cont,column=41).value = fecha.nombre.tecnicaTecnologica
        ws.cell(row=cont,column=42).value = fecha.nombre.acreditacionCarrera
        ws.cell(row=cont,column=43).value = fecha.nombre.cicloEducacion
        ws.cell(row=cont,column=44).value = fecha.nombre.semestreEnCurso
        ws.cell(row=cont,column=45).value = fecha.nombre.modalidad
        ws.cell(row=cont,column=46).value = fecha.nombre.modalidadOtro
        ws.cell(row=cont,column=47).value = fecha.nombre.propedeutico
        ws.cell(row=cont,column=48).value = fecha.nombre.propedeuticoOtro
        ws.cell(row=cont,column=49).value = fecha.nombre.valorMatricula
        ws.cell(row=cont,column=50).value = fecha.nombre.beneficiarioprogramas
        ws.cell(row=cont,column=51).value = fecha.nombre.programaOtro
        ws.cell(row=cont,column=52).value = fecha.nombre.incentivoEducativo
        ws.cell(row=cont,column=53).value = fecha.nombre.cuantasVeces
        ws.cell(row=cont,column=54).value = fecha.nombre.extraAcademicas
        ws.cell(row=cont,column=55).value = fecha.nombre.extraAcademicaPaga
        ws.cell(row=cont,column=56).value = fecha.nombre.parentesco1
        ws.cell(row=cont,column=57).value = fecha.nombre.parentesco2
        ws.cell(row=cont,column=58).value = fecha.nombre.parentesco3
        ws.cell(row=cont,column=59).value = fecha.nombre.parentesco4
        ws.cell(row=cont,column=60).value = fecha.nombre.parentesco5
        ws.cell(row=cont,column=61).value = fecha.nombre.parentesco6
        ws.cell(row=cont,column=62).value = fecha.nombre.parentesco7
        cont = cont + 1
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        #Establecemos el nombre del archivo
    nombre_archivo ="ReporteAgendadosExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

@login_required(login_url = "/login")
def informeAlumnos(request):
    alumnos = Alumno2.objects.all()
    wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'

        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda

        #Creamos los encabezados desde la celda B3 hasta la E3

    ws['b1'] = 'PROCEDENCIA'
    ws['c1'] = 'INSTITUCION'
    ws['d1'] = 'SEDE'
    ws['e1'] = 'NOMBRES'
    ws['f1'] = 'APELLIDOS'
    ws['g1'] = 'DOCUMENTO'
    ws['h1'] = 'GRADO'
    ws['i1'] = 'GENERO'
    ws['j1'] = 'TRANSPORTE'
    cont=4
    #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for alumno in alumnos:
        ws.cell(row=cont,column=2).value = alumno.procedencia.vereda
        ws.cell(row=cont,column=3).value = alumno.institucion.nombre
        ws.cell(row=cont,column=4).value = alumno.sede.sede
        ws.cell(row=cont,column=5).value = alumno.nombres
        ws.cell(row=cont,column=6).value = alumno.apellidos
        ws.cell(row=cont,column=7).value = alumno.documento
        ws.cell(row=cont,column=8).value = alumno.grado.grado
        ws.cell(row=cont,column=9).value = alumno.genero
        ws.cell(row=cont,column=10).value = alumno.transporte
        cont = cont + 1

        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas

        #Establecemos el nombre del archivo
    nombre_archivo ="ReporteAgendaEntrevistaExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

@login_required(login_url = "/login")
def informe(request):
    return render_to_response('informe.html',context_instance = RequestContext(request))

def validaDocumento(request):
    sisben = Sisben.objects.all()
    estrato = Estrato.objects.all()
    municipio = Municipios.objects.all().order_by()
    universidad = Universidades.objects.all()
    vereda = Veredas.objects.all()
    colegio = Colegios.objects.all()
    sedes = Sedes.objects.all()
    centrosSena = CentrosSena.objects.all()
    if request.method == 'POST':
        documento = request.POST['busqueda'] #campo de busqueda de documento de estudiantes
        aspirante = FormularioEstimulos.objects.filter(documento = documento)
        if (len(aspirante)>=1): #si ya se inscribio en estimulos
            fechas = Fechas.objects.all() #fechas de agendamiento
            fechasAgendadas = []
            for f in fechas:
                fechasAgendadas.append(f.nombre.id)
            if (aspirante[0].id in fechasAgendadas): #validar si aspirante ya esta agendado
                nombreAgendado = aspirante[0].nombre
                documen = aspirante[0].documento
                return render_to_response('formulario_exitoso.html',{'nombreAgendado' :nombreAgendado, 'documen':documen}, context_instance=RequestContext(request))
            #aspiranteEstimulo = FormularioEstimulos.objects.get(id = aspirante[0].id )
            #import pdb; pdb.set_trace()
            #return render_to_response('editarFormulario.html',{'aspiranteEstimulo':aspiranteEstimulo, 'municipio':municipio, 'universidad':universidad, 'vereda':vereda, 'colegio':colegio, 'sedes':sedes, 'centrosSena':centrosSena }, context_instance=RequestContext(request))
            else: #sino esta gendado enviar a formulario de estimulos
                aspiranteEstimulo = FormularioEstimulos.objects.get(id = aspirante )
                return render_to_response('resultadoBusquedaUsuario.html',{'aspiranteEstimulo':aspiranteEstimulo, 'municipio':municipio, 'universidad':universidad, 'vereda':vereda, 'colegio':colegio, 'sedes':sedes, 'centrosSena':centrosSena }, context_instance=RequestContext(request))

            #return render_to_response('estimulos_form.html',{'documento' :documento}, context_instance=RequestContext(request))
        else:
            return render_to_response('estimulos_form.html',{'municipio':municipio, 'universidad':universidad, 'vereda':vereda, 'colegio':colegio, 'sedes':sedes, 'centrosSena':centrosSena,'documento':documento}, context_instance = RequestContext(request))
    return render_to_response('informe.html',context_instance = RequestContext(request))

def horaEntrevista(request):
    documento = request.POST['docum']
    if request.method == 'POST':
        hora = Horas.objects.all()
        fecha = request.POST['diaMes']
        estudiante = FormularioEstimulos.objects.filter(documento = documento)
        cita = Fechas.objects.filter(fecha = fecha)
        agendados = Fechas.objects.all()
        listAgendados = []
        for age in agendados:
            listAgendados.append(age.nombre_id)
        if ( estudiante[0].id in listAgendados ):
            nombreAgendado = estudiante[0].nombre
            return render_to_response('formulario_exitoso.html',{'nombreAgendado':nombreAgendado}, context_instance = RequestContext(request))
        if ('hora' in request.POST):
            fechas = Fechas()
            fechas.hora_id = request.POST['hora']
            fechas.fecha = request.POST['diaMes']
            fechas.nombre_id = estudiante[0].id
            fechas.save()
            fechaGuarda = Fechas.objects.filter(nombre = estudiante[0].id)
            nombreGuardado = fechaGuarda[0].nombre.nombre
            diaGuardado = fechaGuarda[0].fecha
            horaGuardada = fechaGuarda[0].hora.hora

        return render_to_response('formulario_exitoso.html',{'nombreGuardado':nombreGuardado, 'diaGuardado':diaGuardado, 'horaGuardada':horaGuardada, 'documen':documento },context_instance = RequestContext(request))

@csrf_exempt
def fechaEntrevista(request):
    documento = request.POST['doc']
    if request.method == 'POST':
        hora = Horas.objects.all()
        fecha = request.POST['dia']
        if fecha != '2018-02-18':
            cita = Fechas.objects.filter(fecha = fecha)
            horasOcupadas = []
            horasOcupadasDisp = []
            for h in range(len(hora)):
                cadaHora = hora[h].id
                for c in range(len(cita)):
                    if (cadaHora==cita[c].hora_id):
                        horasOcupadas.append(cadaHora)
            for elemento in horasOcupadas:
                if horasOcupadas.count(elemento) > 2:
                    horasOcupadasDisp.append(elemento)
            lista_nueva = []
            for i in hora:
                x = i.id
                if (x not in horasOcupadasDisp):
                    lista_nueva.append(i)
            print (horasOcupadas)
            print("--------------------------")
            print (horasOcupadasDisp)
            return render_to_response('formulario_exitoso.html',{ "lista_nueva": lista_nueva, "documento":documento, 'fecha':fecha }, context_instance = RequestContext(request))
        return render_to_response('formulario_exitoso.html',{"documento": documento}, context_instance = RequestContext(request))
    return render_to_response('formulario_exitoso.html', context_instance = RequestContext(request))

@csrf_exempt
def editaFechaEntrevista(request):
    documento = request.POST['doc']
    pk = request.POST['agendado']
    hora = Horas.objects.all()
    fecha = request.POST['dia']
    if fecha != '2018-02-18':
        cita = Fechas.objects.filter(fecha = fecha)
        horasOcupadas = []
        horasOcupadasDisp = []
        for h in range(len(hora)):
            cadaHora = hora[h].id
            for c in range(len(cita)):
                if (cadaHora==cita[c].hora_id):
                    horasOcupadas.append(cadaHora)
        for elemento in horasOcupadas:
            if horasOcupadas.count(elemento) > 2:
                horasOcupadasDisp.append(elemento)
        lista_nueva = []
        for i in hora:
            x = i.id
            if (x not in horasOcupadasDisp):
                lista_nueva.append(i)
        print (horasOcupadas)
        print("--------------------------")
        print (horasOcupadasDisp)
        return render_to_response('editaAgenda.html',{ "lista_nueva": lista_nueva,"pk":pk, "documento":documento, 'fecha':fecha }, context_instance = RequestContext(request))
    return render_to_response('editaAgenda.html',{"documento": documento}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def estimulos(request):
    sisben = Sisben.objects.all()
    estrato = Estrato.objects.all()
    municipio = Municipios.objects.all().order_by()
    universidad = Universidades.objects.all()
    vereda = Veredas.objects.all()
    colegio = Colegios.objects.all()
    sedes = Sedes.objects.all()
    centrosSena = CentrosSena.objects.all()
    video = get_object_or_404(Video, pk=1)
    codigo = 1000
    if request.method == 'POST':
        documento = request.POST['4Documento']
        aspirante = FormularioEstimulos.objects.filter(documento = documento)
        if (len(aspirante)>=1):
            return render_to_response('formulario_exitoso.html',{'documento' :documento}, context_instance=RequestContext(request))
        else:
            formularioEstimulos = FormularioEstimulos()
            formularioEstimulos.nombre = request.POST['1Nombres']
            formularioEstimulos.apellido = request.POST['2Apellidos']
            formularioEstimulos.tipoDocumento = request.POST['3Tipo_documento']
            formularioEstimulos.documentoOtro = request.POST['documentoOtro']
            formularioEstimulos.documento = request.POST['4Documento']
            formularioEstimulos.email = request.POST['5Email']
            formularioEstimulos.celular = request.POST['6Celular']
            formularioEstimulos.genero = request.POST['7Genero']
            formularioEstimulos.edad = request.POST['8Edad']
            formularioEstimulos.estadoCivil = request.POST['9Estado_civil']
            formularioEstimulos.cabezaFamilia = request.POST['10Cabeza_familia']
            formularioEstimulos.hijos = request.POST['hijos']
            formularioEstimulos.pertenencia = request.POST['11pertenencia']
            formularioEstimulos.pertenenciaOtro = request.POST['pertenenciaOtro']
            formularioEstimulos.direccionEstudiante = request.POST['12DireEstudiante']
            formularioEstimulos.ciudadEstudiante = request.POST['Ciudad_estudiante']
            formularioEstimulos.barrioEstudiante = request.POST['Barrio_vereda_estudiante']
            formularioEstimulos.ubicacionEstudiante = request.POST['Ubicacion_residencia_estudiante']
            formularioEstimulos.vivienda = request.POST['13Vivienda']
            formularioEstimulos.estrato = request.POST['14Estrato']
            formularioEstimulos.sisben = request.POST['15SISBEN']
            formularioEstimulos.nucleoFamiliar = request.POST['16Nucleo_familiar']
            formularioEstimulos.direccionPadres = request.POST['17DirePadres']
            formularioEstimulos.ciudadPadres = request.POST['Ciudad_padres']
            formularioEstimulos.barrioPadres = request.POST['Barrio_vereda_padres']
            formularioEstimulos.ubicacionPadres = request.POST['Ubicacion_padres']
            formularioEstimulos.serviciosPublicos = request.POST.getlist('18servicios_publicos[]')
            formularioEstimulos.dependenciaEconomica = request.POST['19dependencia_economica']
            formularioEstimulos.dependenciaOtro = request.POST['dependenciaOtro']
            formularioEstimulos.ingresosFamilia = request.POST['20Ingresos_familia']
            formularioEstimulos.colegio = request.POST['21colegio']
            formularioEstimulos.oficialPrivado = request.POST['22Oficial_privado']
            formularioEstimulos.anoTerminacion = request.POST['23ano_terminacion']
            formularioEstimulos.actaGrado = request.POST['actaGrado']
            formularioEstimulos.universidad = request.POST['24Universidad']
            formularioEstimulos.centrosSena = request.POST['centrosSena']

            formularioEstimulos.sede = request.POST['25Sede']
            formularioEstimulos.acreditacionUniversidad = request.POST['26AcreditaUniversidad']
            formularioEstimulos.carrera = request.POST['27Carrera']
            formularioEstimulos.tecnicaTecnologica = request.POST['tecnicaTecnologica']
            formularioEstimulos.acreditacionCarrera = request.POST['28AcreditaCarrera']
            formularioEstimulos.cicloEducacion = request.POST['29ciclo_educacion']
            formularioEstimulos.semestreEnCurso = request.POST['semestreEnCurso']
            formularioEstimulos.modalidad = request.POST['30Modalidad']
            formularioEstimulos.modalidadOtro = request.POST['modalidadOtro']
            formularioEstimulos.propedeutico = request.POST['31propedeuticos']
            formularioEstimulos.propedeuticoOtro = request.POST['propedeuticoOtro']
            formularioEstimulos.valorMatricula = request.POST['32Valor_matricula']
            formularioEstimulos.beneficiarioprogramas = request.POST['33beneficiario_programas']
            formularioEstimulos.programaOtro = request.POST['programaOtro']
            formularioEstimulos.incentivoEducativo = request.POST['34incentivo_educativo']
            formularioEstimulos.cuantasVeces = request.POST['cuantasVeces']
            formularioEstimulos.extraAcademicas = request.POST['35extra-academicas']
            formularioEstimulos.extraAcademicaPaga = request.POST['36extra_academicas_pagan']
            formularioEstimulos.parentesco1 = request.POST.getlist('parentesco1')
            formularioEstimulos.parentesco2 = request.POST.getlist('parentesco2')
            formularioEstimulos.parentesco3 = request.POST.getlist('parentesco3')
            formularioEstimulos.parentesco4 = request.POST.getlist('parentesco4')
            formularioEstimulos.parentesco5 = request.POST.getlist('parentesco5')
            formularioEstimulos.parentesco6 = request.POST.getlist('parentesco6')
            formularioEstimulos.parentesco7 = request.POST.getlist('parentesco7')
            formularioEstimulos.totalObtenido = 0
            formularioEstimulos.save()
            documento = request.POST['4Documento']
            nombre = request.POST['1Nombres']
            return render_to_response('formulario_exitoso.html',{'codigo':codigo, 'documento':documento, 'nombre':nombre}, context_instance = RequestContext(request))
    return render_to_response('plataformacerrada.html',{'municipio':municipio, 'universidad':universidad, 'vereda':vereda, 'colegio':colegio, 'sedes':sedes, 'centrosSena':centrosSena, 'video':video}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def docentes(request):

    if request.method == 'POST':
        validators = Validator(request.POST)
        validators.required = ['nombre','apellido','documento','email']
        if validators.is_valid():
            docentes = docente()
            docentes.nombres = request.POST['nombre']
            docentes.apellidos = request.POST['apellido']
            docentes.documento = request.POST['documento']
            docentes.email = request.POST['email']
            docentes.asignatura = request.POST['asignatura']
            docentes.contacto = request.POST['contacto']
            docentes.save()
            print(docentes.nombres)
        else:
            return render_to_response('docentes.html', {'error': validator.getMessage() } , context_instance = RequestContext(request))
    return render_to_response('docentes.html',context_instance = RequestContext(request))

@login_required(login_url = "/login")
def matriculas(request):
    grad = Grado.objects.all()
    grupo = Grupos2.objects.all()
    estrato = Estrato.objects.all()
    instituciones = Inst_educativas2.objects.all()
    sedes = SedeInstitucion2.objects.all()
    veredas = Veredas.objects.all()

    if request.method == 'POST':
        #validators = Validator(request.POST)
        #validators.required = ['nombre','apellido','Documento','7Genero','7Genero','15SISBEN','14Estrato','trasporte','12DireEstudiante','familias','alimentacion','desplazado','contacto','dia','email','credo']
        #if validators.is_valid():
        alumno = Alumno2()
        alumno.procedencia_id = request.POST['procedencia']
        alumno.institucion_id = request.POST['institucion']
        alumno.sede_id = request.POST['sede']
        alumno.nombres = request.POST['nombre']
        alumno.apellidos = request.POST['apellido']
        alumno.documento = request.POST['Documento']
        alumno.grado_id = request.POST['grado']
        alumno.genero= request.POST['7Genero']
        alumno.sisben = request.POST['15SISBEN']
        alumno.estrato_id = request.POST['14Estrato']
        alumno.transporte = request.POST['trasporte']
        alumno.direccion = request.POST['12DireEstudiante']
        alumno.familias = request.POST['familias']
        alumno.alimentacion = request.POST['alimentacion']
        alumno.desplazado = request.POST['desplazado']
        alumno.contacto = request.POST['6Celular']
        alumno.nacimiento = request.POST['dia']
        alumno.email = request.POST['email']
        alumno.credo = request.POST['credo']
        alumno.save()
        print(alumno.nombres)
    return render_to_response('matriculas.html', {'gradoss':grad,'estrato':estrato, 'instituciones':instituciones, 'sedes':sedes, 'grupo':grupo, 'veredas':veredas } ,context_instance = RequestContext(request))

@login_required(login_url = "/login")
def actualiza_alumnos(request):
    import pdb; pdb.set_trace()
    pk = request.POST['pk']
    alumno = get_object_or_404(Alumno, pk=pk)
    if request.method == 'POST':
        validators = Validator(request.POST)
        validators.required = ['nombre','apellido','documento','grado','genero','sisben','estrato','trasporte','direccion','familias','alimentacion','desplazado','contacto','nacimiento','email','credo']
        if validators.is_valid():
            alumno.nombres = request.POST['nombre']
            alumno.apellidos = request.POST['apellido']
            alumno.documento = request.POST['documento']
            alumno.grado_id = request.POST['grado']
            alumno.genero= request.POST['genero']
            alumno.sisben = request.POST['sisben']
            alumno.estrato_id = request.POST['estrato']
            alumno.transporte = request.POST['trasporte']
            alumno.direccion = request.POST['direccion']
            alumno.familias = request.POST['familias']
            alumno.alimentacion = request.POST['alimentacion']
            alumno.desplazado = request.POST['desplazado']
            alumno.contacto = request.POST['contacto']
            alumno.nacimiento = request.POST['nacimiento']
            alumno.email = request.POST['email']
            alumno.credo = request.POST['credo']
            alumno.save()
            return render_to_response('busquedas_alumnos.html',context_instance=RequestContext(request))

@login_required(login_url = "/login")
def portada(request):
    if 'matricula' in request.GET.keys():
        return render_to_response('index.html',context_instance = RequestContext(request))
    if 'usuario' in request.GET.keys():
        return render_to_response('formulario.html',context_instance = RequestContext(request))
    if 'doce' in request.GET.keys():
        return render_to_response('docentes.html',context_instance = RequestContext(request))
    return render_to_response('portada.html',context_instance = RequestContext(request))

@login_required(login_url = "/login")
def busquedas(request):
    return render_to_response('busquedas.html',context_instance = RequestContext(request))

@login_required(login_url = "/login")
def esti_form(request):
    return render_to_response('inicio.html',context_instance = RequestContext(request))

def salidaPdf(request):
     def funcion( * args, ** kwargs):
         html = f ( * args, ** kwargs)
         result = StringIO() #creamos una instancia del un objeto StringIO para
         pdf = pisa.pisaDocument( html , result) # convertimos en pdf la template
         return HttpResponse(result.getvalue(), content_type='application/pdf')
     return funcion

@login_required(login_url = "/login")
#@salidaPdf
def busqueda_docente(request):
    docent = None
    buscar = None
    if 'captar' in request.GET.keys():
        buscar = request.GET
        ['dato']
        #qset = (Q(documento__icontains=buscar) )
        qset = ( Q( documento__icontains = buscar) |
                Q( nombres__icontains = buscar) |
                Q( apellidos__icontains = buscar)
                 )
        docent = docente.objects.filter(qset).first()
        print(buscar)
        return render_to_response('busqueda_docente.html', {'docente': docent, 'filtro': buscar},context_instance=RequestContext(request))

@login_required(login_url = "/login")
def busquedas_alumnos(request):
        if request.method == 'POST':
            identificacion = request.POST['busqueda']
            aspirante = FormularioEstimulos.objects.filter(documento = identificacion)
            if (len(aspirante)>=1):
                return render_to_response('resultadobusqueda.html',{'aspirante':aspirante}, context_instance=RequestContext(request))
            alumno = Alumno2.objects.filter(documento = identificacion)
            if (len(alumno)>=1):
                return render_to_response('resultadobusqueda.html',{'alumno':alumno}, context_instance=RequestContext(request))
            else:
                return render_to_response('noresultados.html',context_instance=RequestContext(request))

@login_required(login_url = "/login")
def editarBusqueda(request, pk):
    municipio = Municipios.objects.all().order_by()
    universidad = Universidades.objects.all()
    vereda = Veredas.objects.all()
    colegio = Colegios.objects.all()
    sedes = Sedes.objects.all()
    centrosSena = CentrosSena.objects.all()
    aspiranteEstimulo = FormularioEstimulos.objects.get(id = pk )
    if request.method == 'POST':
        formulario = get_object_or_404(FormularioEstimulos, pk=pk)
        formulario.nombre = request.POST['1Nombres']
        formulario.apellido = request.POST['2Apellidos']
        formulario.tipoDocumento = request.POST['3Tipo_documento']
        formulario.documentoOtro = request.POST['documentoOtro']
        formulario.documento = request.POST['4Documento']
        formulario.email = request.POST['5Email']
        formulario.celular = request.POST['6Celular']
        formulario.genero = request.POST['7Genero']
        formulario.edad = request.POST['8Edad']
        formulario.estadoCivil = request.POST['9Estado_civil']
        formulario.cabezaFamilia = request.POST['10Cabeza_familia']
        formulario.hijos = request.POST['hijos']
        formulario.pertenencia = request.POST['11pertenencia']
        formulario.pertenenciaOtro = request.POST['pertenenciaOtro']
        formulario.direccionEstudiante = request.POST['12DireEstudiante']
        formulario.ciudadEstudiante = request.POST['Ciudad_estudiante']
        formulario.barrioEstudiante = request.POST['Barrio_vereda_estudiante']
        formulario.ubicacionEstudiante = request.POST['Ubicacion_residencia_estudiante']
        formulario.vivienda = request.POST['13Vivienda']
        formulario.estrato = request.POST['14Estrato']
        formulario.sisben = request.POST['15SISBEN']
        formulario.nucleoFamiliar = request.POST['16Nucleo_familiar']
        formulario.direccionPadres = request.POST['17DirePadres']
        formulario.ciudadPadres = request.POST['Ciudad_padres']
        formulario.barrioPadres = request.POST['Barrio_vereda_padres']
        formulario.ubicacionPadres = request.POST['Ubicacion_padres']
        formulario.dependenciaEconomica = request.POST['19dependencia_economica']
        formulario.dependenciaOtro = request.POST['dependenciaOtro']
        formulario.ingresosFamilia = request.POST['20Ingresos_familia']
        formulario.colegio = request.POST['21colegio']
        formulario.oficialPrivado = request.POST['22Oficial_privado']
        formulario.anoTerminacion = request.POST['23ano_terminacion']
        formulario.actaGrado = request.POST['actaGrado']
        formulario.universidad = request.POST['24Universidad']
        formulario.centrosSena = request.POST['centrosSena']

        formulario.sede = request.POST['25Sede']
        formulario.acreditacionUniversidad = request.POST['26AcreditaUniversidad']
        formulario.carrera = request.POST['27Carrera']
        formulario.tecnicaTecnologica = request.POST['tecnicaTecnologica']
        formulario.acreditacionCarrera = request.POST['28AcreditaCarrera']
        formulario.cicloEducacion = request.POST['29ciclo_educacion']
        formulario.semestreEnCurso = request.POST['semestreEnCurso']
        formulario.modalidad = request.POST['30Modalidad']
        formulario.modalidadOtro = request.POST['modalidadOtro']
        formulario.propedeutico = request.POST['31propedeuticos']
        formulario.propedeuticoOtro = request.POST['propedeuticoOtro']
        formulario.valorMatricula = request.POST['32Valor_matricula']
        formulario.beneficiarioprogramas = request.POST['33beneficiario_programas']
        formulario.programaOtro = request.POST['programaOtro']
        formulario.incentivoEducativo = request.POST['34incentivo_educativo']
        formulario.cuantasVeces = request.POST['cuantasVeces']
        formulario.extraAcademicas = request.POST['35extra-academicas']
        formulario.extraAcademicaPaga = request.POST['36extra_academicas_pagan']
        formulario.parentesco1 = request.POST.getlist('parentesco1')
        formulario.parentesco2 = request.POST.getlist('parentesco2')
        formulario.parentesco3 = request.POST.getlist('parentesco3')
        formulario.parentesco4 = request.POST.getlist('parentesco4')
        formulario.parentesco5 = request.POST.getlist('parentesco5')
        formulario.parentesco6 = request.POST.getlist('parentesco6')
        formulario.parentesco7 = request.POST.getlist('parentesco7')
        formulario.totalObtenido = request.POST.getlist('puntajeObtenido')
        formulario.save()
        documento = request.POST['4Documento']
        trazabilidad(request, documento)
        nombre = request.POST['1Nombres']
        nombreAgendado = request.POST['1Nombres']
        documento = documento
        return render_to_response('formulario_exitoso.html',{'nombreAgendado' :nombreAgendado, 'documen':documento}, context_instance=RequestContext(request))

    return render_to_response('editarFormulario.html',{'aspiranteEstimulo':aspiranteEstimulo, 'municipio':municipio, 'universidad':universidad, 'vereda':vereda, 'colegio':colegio, 'sedes':sedes, 'centrosSena':centrosSena }, context_instance=RequestContext(request))

@login_required(login_url = "/login")
def busquedasIncentivos(request):
    estrato = Estrato.objects.filter()
    grado = Grado.objects.filter()
    if request.method == 'POST':
        buscar=None
        alum=None
        buscar=request.POST['busqueda']
        #qset = (Q(documento__icontains=buscar) )
        qset = ( Q( documento = buscar) |
                Q( nombres = buscar) |
                Q( apellidos = buscar)
                 )
        alum = Alumno.objects.filter(qset).first()
        print (alum)
    return render_to_response('formulario_exitoso.html',context_instance=RequestContext(request))

@login_required(login_url = "/login")
def editaBusqAgenda(request):
    horas = Horas.objects.all()
    if request.method == 'POST':
        pk = request.POST['agendado']
        documento = request.POST['documento']
        if request.POST.get('hora'):
            fecha = get_object_or_404(Fechas, pk=pk)
            fecha.fecha = request.POST['diaMes']
            fecha.hora_id = request.POST['hora']
            fecha.nombre_id = request.POST['documento']
            fecha.save()
            nombreGuardado = fecha.nombre.nombre
            diaGuardado = fecha.fecha
            horaGuardada = fecha.hora.hora
            documen =  fecha.nombre.documento
            return render_to_response('agendaExitosa.html',{'documen':documen, 'nombreGuardado':nombreGuardado, 'diaGuardado':diaGuardado, 'horaGuardada':horaGuardada}, context_instance = RequestContext(request))
        return render_to_response('editaAgenda.html',{'horas':horas, 'pk':pk, 'documento':documento}, context_instance = RequestContext(request))
    return render_to_response('editaAgenda.html', context_instance = RequestContext(request))

@login_required(login_url = "/login")
def configuracion(request):
     cita = Fechas.objects.all()
     if request.method == 'POST':
         if 'actualiza' in request._post.keys():
            docentes = docente.objects.get(id=request.user.id)
            docentes.nombres = request.POST['nombre']
            docentes.apellidos = request.POST['apellido']
            docentes.documento = request.POST['documento']
            docentes.email = request.POST['email']
            docentes.asignatura = request.POST['asignatura']
            docentes.contacto = request.POST['contacto']
            docentes.save()
            return render_to_response('actualiza_docentes.html',{'docente':docentes},context_instance = RequestContext(request))
     return render_to_response('configuracionEstimulos.html',{'cita':cita}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def logout(request):
    auth.logout(request)
    return HttpResponseRedirect("/login")

from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import Http404

@login_required(login_url = "/login")
def municipios(request):
    municipios = Municipios.objects.all()
    if request.method == 'POST':
        municipio = Municipios()
        municipio.municipio = request.POST['municipio']
        municipio.save()
        return render_to_response('municipios.html',{'municipios':municipios}, context_instance = RequestContext(request))
    return render_to_response('municipios.html',{'municipios':municipios}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def editarMunicipio(request, pk):
    municipio = get_object_or_404(Municipios, pk=pk)
    if request.method == 'POST':
        municipio = get_object_or_404(Municipios, pk=pk)
        municipio.municipio = request.POST['municipio']
        municipio.save()
        municipios = Municipios.objects.all()
        return render_to_response('municipios.html',{'municipios':municipios}, context_instance = RequestContext(request))
    return render_to_response('editarMunicipio.html',{'municipio':municipio}, context_instance = RequestContext(request))

@permission_required('Municipios.can_delete_Municipios', login_url = '/login')
def eliminarMunicipio(request, pk):
    municipio = get_object_or_404(Municipios, pk=pk)
    if request.method == 'POST':
        municipio.delete()
        municipios = Municipios.objects.all()
        return render_to_response('municipios.html',{'municipios':municipios}, context_instance = RequestContext(request))
    return render_to_response('eliminarMunicipio.html',{'municipio':municipio}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def veredas(request):
    veredas = Veredas.objects.all()
    if request.method == 'POST':
        vereda = Veredas()
        vereda.vereda = request.POST['vereda']
        vereda.save()
        return render_to_response('veredas.html',{'veredas':veredas}, context_instance = RequestContext(request))
    return render_to_response('veredas.html',{'veredas':veredas}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def editarVereda(request, pk):
    vereda = get_object_or_404(Veredas, pk=pk)
    if request.method == 'POST':
        vereda = get_object_or_404(Veredas, pk=pk)
        vereda.vereda = request.POST['vereda']
        vereda.save()
        veredas = Veredas.objects.all()
        return render_to_response('veredas.html',{'veredas':veredas}, context_instance = RequestContext(request))
    return render_to_response('editarVereda.html',{'vereda':vereda}, context_instance = RequestContext(request))

@permission_required('Veredas.can_delete_Veredas', login_url = '/login')
def eliminarVereda(request, pk):
    vereda = get_object_or_404(Veredas, pk=pk)
    if request.method == 'POST':
        vereda.delete()
        veredas = Veredas.objects.all()
        return render_to_response('veredas.html',{'veredas':veredas}, context_instance = RequestContext(request))
    return render_to_response('eliminarVereda.html',{'vereda':vereda}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def universidades(request):
    universidades = Universidades.objects.all()
    if request.method == 'POST':
        universidades = Universidades()
        universidades.universidad = request.POST['universidad']
        universidades.save()
        return render_to_response('universidades.html',{'universidades':universidades}, context_instance = RequestContext(request))
    return render_to_response('universidades.html',{'universidades':universidades},context_instance = RequestContext(request))

@login_required(login_url = "/login")
def editarUniversidad(request, pk):
    universidad = get_object_or_404(Universidades, pk=pk)
    if request.method == 'POST':
        universidad = get_object_or_404(Universidades, pk=pk)
        universidad.universidad = request.POST['universidad']
        universidad.save()
        universidades = Universidades.objects.all()
        return render_to_response('universidades.html',{'universidades':universidades}, context_instance = RequestContext(request))
    return render_to_response('editarUniversidad.html',{'universidad':universidad}, context_instance = RequestContext(request))

@permission_required('Universidades.can_delete_Universidades', login_url = '/login')
def eliminarUniversidad(request, pk):
    universidad = get_object_or_404(Universidades, pk=pk)
    if request.method == 'POST':
        universidad.delete()
        universidades = Universidades.objects.all()
        return render_to_response('universidades.html',{'universidades':universidades}, context_instance = RequestContext(request))
    return render_to_response('eliminarUniversidad.html',{'universidad':universidad}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def colegios(request):
    colegios = Colegios.objects.all()
    if request.method == 'POST':
        colegio = Colegios()
        colegio.colegio = request.POST['colegio']
        colegio.save()
        return render_to_response('colegios.html',{'colegios':colegios}, context_instance = RequestContext(request))
    return render_to_response('colegios.html',{'colegios':colegios}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def editarColegio(request, pk):
    colegio = get_object_or_404(Colegios, pk=pk)
    if request.method == 'POST':
        colegio = get_object_or_404(Colegios, pk=pk)
        colegio.colegio = request.POST['colegio']
        colegio.save()
        colegios = Colegios.objects.all()
        return render_to_response('colegios.html',{'colegios':colegios}, context_instance = RequestContext(request))
    return render_to_response('editarColegio.html',{'colegio':colegio}, context_instance = RequestContext(request))

@permission_required('Colegios.can_delete_Colegios', login_url = '/login')
def eliminarColegio(request, pk):
    colegio = get_object_or_404(Colegios, pk=pk)
    if request.method == 'POST':
        colegio.delete()
        colegios = Colegios.objects.all()
        return render_to_response('colegios.html',{'colegios':colegios}, context_instance = RequestContext(request))
    return render_to_response('eliminarColegio.html',{'colegio':colegio}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def centrosSena(request):
    centrosSena = CentrosSena.objects.all()
    if request.method == 'POST':
        centroSena = CentrosSena()
        centroSena.centroSena = request.POST['centroSena']
        centroSena.save()
        return render_to_response('centrosSena.html',{'centrosSena':centrosSena}, context_instance = RequestContext(request))
    return render_to_response('centrosSena.html',{'centrosSena':centrosSena}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def editarCentrosSena(request, pk):
    centroSena = get_object_or_404(CentrosSena, pk=pk)
    if request.method == 'POST':
        centroSena = get_object_or_404(CentrosSena, pk=pk)
        centroSena.centroSena = request.POST['centroSena']
        centroSena.save()
        centrosSena = CentrosSena.objects.all()
        return render_to_response('centrosSena.html',{'centrosSena':centrosSena}, context_instance = RequestContext(request))
    return render_to_response('editarCentroSena.html',{'centroSena':centroSena}, context_instance = RequestContext(request))

@permission_required('CentrosSena.can_delete_CentrosSena', login_url = '/login')
def eliminarCentrosSena(request, pk):
    centroSena = get_object_or_404(CentrosSena, pk=pk)
    if request.method == 'POST':
        centroSena.delete()
        centrosSena = CentrosSena.objects.all()
        return render_to_response('centrosSena.html',{'centrosSena':centrosSena}, context_instance = RequestContext(request))
    return render_to_response('eliminarCentroSena.html',{'centroSena':centroSena}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def sedes(request):
    sedes = Sedes.objects.all()
    if request.method == 'POST':
        sede = Sedes()
        sede.sede = request.POST['sedes']
        sede.save()
        return render_to_response('sedes.html',{'sedes':sedes}, context_instance = RequestContext(request))
    return render_to_response('sedes.html',{'sedes':sedes}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def editarSedeUni(request, pk):
    sede = get_object_or_404(Sedes, pk=pk)
    if request.method == 'POST':
        sede = get_object_or_404(Sedes, pk=pk)
        sede.sede = request.POST['sede']
        sede.save()
        sedes = Sedes.objects.all()
        return render_to_response('sedes.html',{'sedes':sedes}, context_instance = RequestContext(request))
    return render_to_response('editarSedeUniversidad.html',{'sede':sede}, context_instance = RequestContext(request))

@permission_required('Sedes.can_delete_Sedes', login_url = '/login')
def eliminarSedeUni(request, pk):
    sede = get_object_or_404(Sedes, pk=pk)
    if request.method == 'POST':
        sede.delete()
        sedes = Sedes.objects.all()
        return render_to_response('sedes.html',{'sedes':sedes}, context_instance = RequestContext(request))
    return render_to_response('eliminarSedeUniversidad.html',{'sede':sede}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def video(request):
    video = get_object_or_404(Video, pk=1)
    if request.method == 'POST':
        video = get_object_or_404(Video, pk=1)
        video.video = request.POST['video']
        video.save()
        return render_to_response('video.html',{'video':video}, context_instance = RequestContext(request))
    return render_to_response('video.html',{'video':video}, context_instance = RequestContext(request))

def salidaPdf(f):

    def funcion(*args, **kwargs):
        html = f(*args, **kwargs)
        result = StringIO() #creamos una instancia del un objeto StringIO para
        pdf = pisa.pisaDocument( html , result) # convertimos en pdf la template
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return funcion


@salidaPdf
def reporte(request):
    if request.method == 'POST':
        documento = request.POST['doc']
        data = FormularioEstimulos.objects.filter(documento=documento).first()
        return render_to_string('reporteincentivo.html',{'data':data})

@csrf_exempt
def listaEntrevista(request):

    horas = Horas.objects.all()
    if request.method == 'POST':
        if  request.POST['busca'] != "":
            identificacion = request.POST['busca']
            aspirante = FormularioEstimulos.objects.filter(documento = identificacion)
            if (len(aspirante)>=1):
                agendado = aspirante[0].id
                print (agendado)
                agendadoFecha = Fechas.objects.filter(nombre = agendado)
                agendadoDatos = agendadoFecha[0]
                print (agendadoFecha)
                return render_to_response('resultBusqAgenda.html',{'agendadoDatos':agendadoDatos}, context_instance=RequestContext(request))
            else:
                return render_to_response('noresultados.html',context_instance=RequestContext(request))

        if  request.POST['hora'] == "NR":
            fecha = request.POST['fecha']
            listaHoras = []
            estudianteHora = []
            totalHora = []
            for h in range(len(horas)):
                hor = horas[h].id
                cita = Fechas.objects.filter(fecha = fecha, hora = hor)
                for c in range(len(cita)):
                    if c == 0:
                        estudianteHora.append("hora: "+horas[h].hora)
                    estudianteHora.append(cita[c].nombre.nombre+" | "+cita[c].nombre.apellido+" | C.c "+cita[c].nombre.documento+" | Cel: "+cita[c].nombre.celular)
            return render_to_response('entrevistas.html',{'estudianteHora':estudianteHora, 'fecha':fecha, 'horas':horas},context_instance = RequestContext(request))
        else:
            fecha = request.POST['fecha']
            hora = request.POST['hora']
            estudiante = []
            cita = Fechas.objects.filter(fecha = fecha, hora = hora)
            for c in range(len(cita)):
                estudiante.append(cita[c].nombre.nombre+" | "+cita[c].nombre.apellido+" | C.c "+cita[c].nombre.documento+" | Cel: "+cita[c].nombre.celular)
            return render_to_response('entrevistas.html',{'estudiante':estudiante, 'fecha':fecha, 'horas':horas, 'cita':cita},context_instance = RequestContext(request))

    return render_to_response('entrevistas.html',{'horas':horas}, context_instance = RequestContext(request))

@login_required(login_url = "/login")
def trazabilidad(request, action):
    traza = Traza()
    traza.usuario = request.user.id
    if (Usuario.objects.filter(id=request.user.id).exists()):
        traza.tipo = 'E'
    elif (Profesor.objects.filter(id=request.user.id).exists()):
        traza.tipo = 'P'
    else:
        traza.tipo = 'A'
    traza.accion = action
    traza.save()

def editarBusquedaParcialmente(request, pk):
    municipio = Municipios.objects.all().order_by()
    universidad = Universidades.objects.all()
    vereda = Veredas.objects.all()
    colegio = Colegios.objects.all()
    sedes = Sedes.objects.all()
    centrosSena = CentrosSena.objects.all()
    aspiranteEstimulo = FormularioEstimulos.objects.get(id = pk )
    if request.method == 'POST':
        documento = request.POST['4Documento']
        aspirante = FormularioEstimulos.objects.filter(documento = documento)
        if (len(aspirante)>=1): # si existe el aspirante
            fechas = Fechas.objects.all()
            fechasAgendadas = []
            for f in fechas:
                fechasAgendadas.append(f.nombre.id)
            if (aspirante[0].id in fechasAgendadas): # si el aspirante esta agendado
                nombreAgendado = aspirante[0].nombre
                documento = aspirante[0].documento
                return render_to_response('formulario_exitoso.html',{'documento':documento}, context_instance=RequestContext(request))
            else:
                formulario = get_object_or_404(FormularioEstimulos, pk=pk)
                formulario.nombre = request.POST['1Nombres']
                formulario.apellido = request.POST['2Apellidos']
                formulario.tipoDocumento = request.POST['3Tipo_documento']
                formulario.documentoOtro = request.POST['documentoOtro']
                formulario.documento = request.POST['4Documento']
                formulario.celular = request.POST['6Celular']
                formulario.edad = request.POST['8Edad']
                formulario.direccionEstudiante = request.POST['12DireEstudiante']
                formulario.ciudadEstudiante = request.POST['Ciudad_estudiante']
                formulario.barrioEstudiante = request.POST['Barrio_vereda_estudiante']
                formulario.ubicacionEstudiante = request.POST['Ubicacion_residencia_estudiante']
                formulario.direccionPadres = request.POST['17DirePadres']
                formulario.ciudadPadres = request.POST['Ciudad_padres']
                formulario.barrioPadres = request.POST['Barrio_vereda_padres']
                formulario.ubicacionPadres = request.POST['Ubicacion_padres']
                formulario.semestreEnCurso = request.POST['semestreEnCurso']
                formulario.valorMatricula = request.POST['32Valor_matricula']
                formulario.totalObtenido = request.POST.getlist('puntajeObtenido')
                formulario.save()
                documento = request.POST['4Documento']
                trazabilidad(request, documento)
                nombre = request.POST['1Nombres']
                nombreAgendado = request.POST['1Nombres']
                documen = documento
                return render_to_response('formulario_exitoso.html',{'documento':documento}, context_instance=RequestContext(request))
        return render_to_response('editarFormularioPorUsuario.html',{'aspiranteEstimulo':aspiranteEstimulo, 'municipio':municipio, 'universidad':universidad, 'vereda':vereda, 'colegio':colegio, 'sedes':sedes, 'centrosSena':centrosSena }, context_instance=RequestContext(request))
    return render_to_response('editarFormularioPorUsuario.html',{'aspiranteEstimulo':aspiranteEstimulo, 'municipio':municipio, 'universidad':universidad, 'vereda':vereda, 'colegio':colegio, 'sedes':sedes, 'centrosSena':centrosSena }, context_instance=RequestContext(request))


def editarFormularioUsuario(request):
    return render_to_response('editarFormularioPorUsuario.html',context_instance = RequestContext(request))
